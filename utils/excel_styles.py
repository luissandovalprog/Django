"""
Utilidades para crear archivos Excel con diseño profesional usando openpyxl
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def aplicar_estilo_encabezado(ws, row=1):
    """
    Aplica estilo profesional a la fila de encabezados
    
    Args:
        ws: Worksheet de openpyxl
        row: Número de fila (por defecto 1)
    """
    # Estilo de encabezado: Fondo verde, texto blanco, negrita
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def aplicar_bordes(ws, max_row, max_col):
    """
    Aplica bordes a todas las celdas con datos
    
    Args:
        ws: Worksheet de openpyxl
        max_row: Número máximo de filas
        max_col: Número máximo de columnas
    """
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def ajustar_ancho_columnas(ws, max_col, min_width=10, max_width=50):
    """
    Ajusta automáticamente el ancho de las columnas basado en el contenido
    
    Args:
        ws: Worksheet de openpyxl
        max_col: Número máximo de columnas
        min_width: Ancho mínimo de columna (por defecto 10)
        max_width: Ancho máximo de columna (por defecto 50)
    """
    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Ajustar ancho con límites
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def aplicar_filtros_automaticos(ws, max_col):
    """
    Activa filtros automáticos en la fila de encabezados
    
    Args:
        ws: Worksheet de openpyxl
        max_col: Número máximo de columnas
    """
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"


def congelar_paneles(ws, row=2, col=1):
    """
    Congela los paneles para mantener encabezados visibles
    
    Args:
        ws: Worksheet de openpyxl
        row: Fila a partir de la cual congelar (por defecto 2 para congelar fila 1)
        col: Columna a partir de la cual congelar (por defecto 1)
    """
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def aplicar_estilo_datos_alternados(ws, max_row, max_col, start_row=2):
    """
    Aplica color de fondo alternado a las filas de datos (efecto zebra)
    
    Args:
        ws: Worksheet de openpyxl
        max_row: Número máximo de filas
        max_col: Número máximo de columnas
        start_row: Fila donde empiezan los datos (por defecto 2)
    """
    light_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    for row_idx in range(start_row, max_row + 1):
        # Aplicar color solo a filas pares
        if row_idx % 2 == 0:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = light_fill


def centrar_columna(ws, col_idx, max_row):
    """
    Centra el contenido de una columna específica
    
    Args:
        ws: Worksheet de openpyxl
        col_idx: Índice de la columna (1-indexed)
        max_row: Número máximo de filas
    """
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx in range(2, max_row + 1):  # Empezar desde fila 2 (después del encabezado)
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = center_alignment


def aplicar_formato_numerico(ws, col_idx, max_row, formato='#,##0'):
    """
    Aplica formato numérico a una columna específica
    
    Args:
        ws: Worksheet de openpyxl
        col_idx: Índice de la columna (1-indexed)
        max_row: Número máximo de filas
        formato: String de formato Excel (por defecto separador de miles)
    """
    for row_idx in range(2, max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = formato


def aplicar_estilo_completo(wb, ws, max_row, max_col, columnas_centradas=None, columnas_numericas=None):
    """
    Aplica todos los estilos profesionales a una hoja de Excel
    
    Args:
        wb: Workbook de openpyxl
        ws: Worksheet de openpyxl
        max_row: Número máximo de filas
        max_col: Número máximo de columnas
        columnas_centradas: Lista de índices de columnas a centrar (opcional)
        columnas_numericas: Dict con {col_idx: formato} para columnas numéricas (opcional)
    """
    # Aplicar estilos básicos
    aplicar_estilo_encabezado(ws, row=1)
    aplicar_bordes(ws, max_row, max_col)
    ajustar_ancho_columnas(ws, max_col)
    aplicar_filtros_automaticos(ws, max_col)
    congelar_paneles(ws, row=2, col=1)
    aplicar_estilo_datos_alternados(ws, max_row, max_col, start_row=2)
    
    # Aplicar centrado a columnas específicas
    if columnas_centradas:
        for col_idx in columnas_centradas:
            centrar_columna(ws, col_idx, max_row)
    
    # Aplicar formato numérico a columnas específicas
    if columnas_numericas:
        for col_idx, formato in columnas_numericas.items():
            aplicar_formato_numerico(ws, col_idx, max_row, formato)
